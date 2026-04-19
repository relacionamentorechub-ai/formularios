#!/usr/bin/env python3
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Cores da identidade
COR_HEADER_BG = "1A1A2E"
COR_HEADER_FG = "FFFFFF"
COR_TOP3_BG = "F0C040"
COR_TOP3_FG = "1A1A2E"
COR_LINHA_PAR = "F7F7F7"
COR_LINHA_IMPAR = "FFFFFF"
COR_BORDA = "CCCCCC"
COR_TITULO_BG = "16213E"
COR_INSTAGRAM = "C13584"

FONTE_PRINCIPAL = "Arial"

def thin_border():
    side = Side(style="thin", color=COR_BORDA)
    return Border(left=side, right=side, top=side, bottom=side)

def aplicar_header(ws, row, col, texto, negrito=True, tamanho=11):
    cell = ws.cell(row=row, column=col, value=texto)
    cell.font = Font(
        name=FONTE_PRINCIPAL, bold=negrito, color=COR_HEADER_FG, size=tamanho
    )
    cell.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    return cell

def criar_planilha(dados, output_path):
    cidade = dados.get("cidade", "Cidade")
    locais = dados.get("locais", [])
    top3_idx = set(dados.get("top3", []))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Locais - {cidade[:20]}"

    # Titulo
    ws.merge_cells("A1:G1")
    titulo = ws["A1"]
    titulo.value = f"LOCAIS PARA EVENTOS — {cidade.upper()}"
    titulo.font = Font(name=FONTE_PRINCIPAL, bold=True, color=COR_HEADER_FG, size=14)
    titulo.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtítulo
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = f"Pesquisa realizada em {datetime.now().strftime('%d/%m/%Y')}   |   Destacados em dourado = Top 3 recomendados"
    sub.font = Font(name=FONTE_PRINCIPAL, italic=True, color="555555", size=9)
    sub.fill = PatternFill("solid", fgColor="EEF2FF")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Cabeçalho
    colunas = [
        ("NOME DO LOCAL", 30),
        ("TIPO", 18),
        ("INSTAGRAM", 32),
        ("ENDERECO", 38),
        ("CAPACIDADE", 14),
        ("NIVEL / PUBLICO", 20),
        ("OBSERVACOES", 55),
    ]

    ROW_HEADER = 3
    for col_idx, (nome_col, largura) in enumerate(colunas, start=1):
        aplicar_header(ws, ROW_HEADER, col_idx, nome_col)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[ROW_HEADER].height = 28

    # Dados
    for idx, local in enumerate(locais):
        row = ROW_HEADER + 1 + idx
        is_top3 = idx in top3_idx

        if is_top3:
            bg = COR_TOP3_BG
            fg = COR_TOP3_FG
        else:
            bg = COR_LINHA_PAR if idx % 2 == 0 else COR_LINHA_IMPAR
            fg = "1A1A2E"

        fill = PatternFill("solid", fgColor=bg)
        fonte_base = Font(name=FONTE_PRINCIPAL, color=fg, size=10)
        fonte_negrito = Font(name=FONTE_PRINCIPAL, color=fg, size=10, bold=True)

        def _cell(col, valor, negrito=False, wrap=False, h_align="left"):
            c = ws.cell(row=row, column=col, value=valor)
            c.font = fonte_negrito if negrito else fonte_base
            c.fill = fill
            c.alignment = Alignment(
                horizontal=h_align, vertical="top", wrap_text=wrap
            )
            c.border = thin_border()
            return c

        # Nome
        nome_val = local.get("nome", "")
        if is_top3:
            nome_val = f"* {nome_val}"
        _cell(1, nome_val, negrito=is_top3)

        # Tipo
        _cell(2, local.get("tipo", ""), h_align="center")

        # Instagram
        ig_url = local.get("instagram", "")
        ig_handle = local.get("instagram_handle", ig_url)
        cell_ig = ws.cell(row=row, column=3)
        cell_ig.fill = fill
        cell_ig.border = thin_border()
        cell_ig.alignment = Alignment(horizontal="left", vertical="top")

        if ig_url:
            cell_ig.value = ig_handle
            cell_ig.hyperlink = ig_url
            cell_ig.font = Font(
                name=FONTE_PRINCIPAL,
                color=COR_INSTAGRAM if not is_top3 else "8B0057",
                size=10,
                underline="single",
                bold=is_top3,
            )
        else:
            cell_ig.value = "—"
            cell_ig.font = fonte_base

        # Endereco
        _cell(4, local.get("endereco", ""), wrap=True)

        # Capacidade
        cap = local.get("capacidade", "")
        aval = local.get("avaliacao_google", "")
        cap_texto = cap
        if aval:
            cap_texto = f"{cap}\n{aval}"
        _cell(5, cap_texto, h_align="center", wrap=True)

        # Nivel / Publico
        nivel = local.get("nivel_preco", "")
        publico = local.get("publico_ideal", "")
        _cell(6, f"{nivel}\n{publico}" if nivel or publico else "—", wrap=True, h_align="center")

        # Observacoes
        obs_partes = []
        tipos_ev = local.get("tipos_evento", "")
        if tipos_ev:
            obs_partes.append(f"Eventos: {tipos_ev}")
        fortes = local.get("pontos_fortes", "")
        if fortes:
            obs_partes.append(f"Strengths: {fortes}")
        obs_extra = local.get("observacoes", "")
        if obs_extra:
            obs_partes.append(obs_extra)

        _cell(7, "\n".join(obs_partes), wrap=True)
        ws.row_dimensions[row].height = max(50, 15 * (len(obs_partes) + 1))

    # Legenda
    row_legenda = ROW_HEADER + 1 + len(locais) + 1
    ws.merge_cells(f"A{row_legenda}:G{row_legenda}")
    leg = ws[f"A{row_legenda}"]
    leg.value = "* = Top 3 recomendados  |  Links Instagram clicaveis  |  Pesquisa via Google + Instagram"
    leg.font = Font(name=FONTE_PRINCIPAL, italic=True, color="888888", size=8)
    leg.alignment = Alignment(horizontal="center", vertical="center")
    leg.fill = PatternFill("solid", fgColor="F0F0F0")

    # Congelar
    ws.freeze_panes = f"A{ROW_HEADER + 1}"

    # Salvar
    wb.save(output_path)
    return output_path

# Main
json_file = r"C:\Users\Usuario\Downloads\REC - HUB COMPLETO\REC - HUB COMPLETO\venues_canoas_temp.json"
with open(json_file, encoding="utf-8") as f:
    dados = json.load(f)

cidade = dados.get("cidade", "Canoas")
output_dir = Path(r"C:\Users\Usuario\Downloads\REC - HUB COMPLETO\REC - HUB COMPLETO\CIDADES PARA EVENTOS") / cidade.upper()
output_dir.mkdir(parents=True, exist_ok=True)
output_file = output_dir / f"Locais para Eventos - {cidade.title()}.xlsx"

caminho = criar_planilha(dados, str(output_file))
print(f"Planilha salva em: {caminho}")
print(f"Planilha gerada com {len(dados.get('locais', []))} locais.")
